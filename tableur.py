import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill

import os.path
from pathlib import Path


from .fonction import afficheerreur
from .constante import *


class Tableur:

    def __init__(self):
        # classeur
        self.wbook = None
        # feuille
        self.wsfeuille = None

    def initclasseur(self):
        fichier_excel = Path(os.path.dirname(__file__) + "/transaction.xlsx")
        if not fichier_excel.exists():
            self.wbook = openpyxl.Workbook()
            self.wsfeuille = self.wbook.active
            self.wsfeuille.title = "Contribution directe"
            self.wsfeuille["A1"] = "Date"
            self.wsfeuille["B1"] = "N° de transaction"
            self.wsfeuille["C1"] = CLEABS
            self.wsfeuille["D1"] = NATURE + " avant"
            self.wsfeuille["E1"] = NATURE + " après"
            self.wsfeuille["F1"] = NB_VOIES + " avant"
            self.wsfeuille["G1"] = NB_VOIES + " après"
            self.wsfeuille["H1"] = LARGEUR + " avant"
            self.wsfeuille["I1"] = LARGEUR + " après"
            self.wsfeuille["J1"] = IMPORTANCE + " avant"
            self.wsfeuille["K1"] = IMPORTANCE + " après"
            self.wsfeuille["L1"] = SENS + " avant"
            self.wsfeuille["M1"] = SENS + " après"
            self.wsfeuille["N1"] = ACCES + " avant"
            self.wsfeuille["O1"] = ACCES + " après"

            self.wsfeuille.column_dimensions["A"].width = 25
            self.wsfeuille.column_dimensions["B"].width = 28
            self.wsfeuille.column_dimensions["C"].width = 28
            self.wsfeuille.column_dimensions["D"].width = 30
            self.wsfeuille.column_dimensions["E"].width = 30
            self.wsfeuille.column_dimensions["F"].width = 30
            self.wsfeuille.column_dimensions["G"].width = 30
            self.wsfeuille.column_dimensions["H"].width = 30
            self.wsfeuille.column_dimensions["I"].width = 30
            self.wsfeuille.column_dimensions["J"].width = 30
            self.wsfeuille.column_dimensions["K"].width = 30
            self.wsfeuille.column_dimensions["L"].width = 30
            self.wsfeuille.column_dimensions["M"].width = 30
            self.wsfeuille.column_dimensions["N"].width = 30
            self.wsfeuille.column_dimensions["O"].width = 30

            # Formater les titres (mettre en gras et centrer)
            bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            cellule_entete = ["A1","B1","C1","D1","E1","F1","G1","H1","I1","J1","K1","L1","M1","N1","O1"]
            for cell in cellule_entete:
                self.wsfeuille[cell].font = bold_font
                # self.wsfeuille[cell].alignment = center_alignment

            # Formater les cellules (ex. : bordures autour des données)
            thick_border = Border(left=Side(style="thick"),
                                 right=Side(style="thick"),
                                 top=Side(style="thick"),
                                 bottom=Side(style="thick"))

            # Appliquer les bordures à toutes les cellules contenant des données
            for row in self.wsfeuille.iter_rows(min_row=1, max_row=self.wsfeuille.max_row, min_col=1, max_col=self.wsfeuille.max_column):
                for cell in row:
                    cell.border = thick_border

            # Formater la colonne A (colonne 1) pour qu'Excel reconnaisse bien les dates
            for row in self.wsfeuille.iter_rows(min_row=1, max_row=self.wsfeuille.max_row, min_col=1, max_col=1):
                for cell in row:
                    # Définir le format de la cellule pour afficher les dates au format 'YYYY-MM-DD'
                    cell.number_format = 'YYYY-MM-DD'

        else:
            self.wbook = openpyxl.load_workbook(fichier_excel)
            self.wsfeuille = self.wbook["Contribution directe"]

    def log_is_open(self):
        # on test si le fichier est deja ouvert
        try:
            # Tente d'ouvrir le fichier en mode écriture
            with open(os.path.dirname(__file__) + "/transaction.xlsx", 'r+'):
                # le fichier n'est pas ouvert
                return True

        except PermissionError:
            afficheerreur(
                "La modification n'a pas été appliquée<br>Le fichier de <b>log</b> : <mark style='background-color: "
                "lightgreen;'>transaction.xlsx</mark> est ouvert<p><br>Veuillez le fermer pour pouvoir ecrire de "
                "nouvelles données")
            return False  # Une erreur signifie que le fichier est déjà ouvert
        except FileNotFoundError:
            return True

    def adddonnees(self,listdonnees):
        # on initialise un nouveau classeur s'il n'existe pas
        # sinon on rempli dans le classeur existant
        self.initclasseur()

        cell_color = PatternFill(start_color="ff978d", end_color="ff978d", fill_type="solid")
        couple_cell = [["D2", "E2"], ["F2", "G2"], ["H2", "I2"], ["J2", "K2"], ["L2", "M2"], ["N2", "O2"]]

        for ligne in listdonnees:
            # inserer une ligne vide a la position 2 (premiere ligne apres les titres)
            self.wsfeuille.insert_rows(2)
            # ajout des données dans la ligne crée
            for col, valeur in enumerate(ligne, start=1):
                # afficheerreur(valeur)
                self.wsfeuille.cell(row = 2, column = col, value = valeur)

            # colorier les celulles dont les attributs ont changés
            for couple in couple_cell:
                cellule1 = self.wsfeuille[couple[0]]
                cellule2 = self.wsfeuille[couple[1]]
                if cellule1.value != cellule2.value:
                    # afficheerreur("nature differente , on colorie")
                    cellule1.fill = cell_color
                    cellule2.fill = cell_color


        # centre toutes les valeurs des celulles
        for row in self.wsfeuille.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')


        return True

    def add_colonne(self,colonne):
        pass

    def sauvegarder(self):
        self.wbook.save(os.path.dirname(__file__) + "/transaction.xlsx")


